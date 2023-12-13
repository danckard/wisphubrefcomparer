import PyPDF2
import platform
import re
import os
import openpyxl
from xls2xlsx import XLS2XLSX
from openpyxl.styles import PatternFill
import pyautogui
import ctypes  # An included library with Python install.
import requests
import difflib
import random
from tqdm import tqdm, trange

def progressbar(value,maxvalue):
    pbar = tqdm(total=int(maxvalue),colour='blue')
    pbar.update(value)

# def key(choices, keyword):
#     matches = difflib.get_close_matches(keyword, choices)
#     if matches:
#         best_match, *_ = matches
#         return difflib.SequenceMatcher(None, keyword, best_match).ratio()
#     return 0.0

banescofiles=[]
mercantilfiles=[]
bncfiles=[]
references=[]
reflocations=[]
# creating a pdf file object
def processpdf(filename):
    print('procesando archivo:'+str(filename))
    referencespdf=[]
    reflocations=[]
    banktype=None
    counter=0
    pdfFileObj = open(str(filename), 'rb')
    if re.findall('mercantil',str(((filename).lower()))):
        banktype='mercantil'
    pdfReader = PyPDF2.PdfReader(pdfFileObj)
    for page in pdfReader.pages:
        if banktype=='mercantil':
            lines=(page.extract_text()).split('\n')
            for line in lines:
                if re.search('[0-9][0-9]+/+[0-9][0-9]+/+[0-9][0-9]',line):
                    try:
                        if "Mercantil en Línea" not in line:
                            reference=str((line.split(' '))[1])
                            counter+=1
                    except:
                        continue
                    if "Mercantil en Línea" not in line:
                        # if reference.startswith('000000'):
                        #     # reference=cellvalue[6:]
                        #     print('detectado ref provincial')
                        # if reference.startswith('0000'):
                        #     print('detectado ref provincial')
                        if len(reference)>=7:
                            reference=reference[4:]
                        # print(reference)
                        referencespdf.append(str(reference))
                        reflocations.append(str(reference)+':'+str(counter)+':'+str(filename))
                
    referencespdf=referencespdf
    return set(referencespdf),set(reflocations)
def processxls(filename):
    print('procesando archivo:'+str(filename))
    referencesexcel=[]
    wb_obj = openpyxl.load_workbook(filename)
    ws = wb_obj.active
    banktype=None
    reflocations=[]
    if re.findall('banesco',str(((filename).lower()))):
        banktype='banesco'
    if re.findall('bnc',str(((filename).lower()))):
        banktype='bnc'
    
    for thenumber,row in enumerate(ws.iter_rows()):
        if platform.system() =="Windows":
            clear = lambda: os.system('cls')
            # clear()
        else:
            clear = lambda: os.system('clear')
            # clear()
        # if random.randint(1,10)>=8:
        progressbar(int(thenumber),int(ws.max_row))
        # time.sleep(0.5)
        for index, cell in enumerate(row):
            if banktype=='banesco':
                refletter='B'
            if banktype=='bnc':
                refletter='M'
            try:
                
                if cell.column_letter==refletter:
                    
                    if re.search('\d+',str((cell.value))):
                        cellvalue=str(cell.value)
                        if cellvalue.startswith('000000'):
                            # print('detectado provincial')
                            cellvalue=cellvalue[5:]
                        if cellvalue.startswith('0000'):
                            print('detectado provincial')
                            cellvalue=cellvalue[4:]
                        # print(cell.value)
                        referencesexcel.append(str(cellvalue))
                        reflocations.append(str(cellvalue)+':'+str(cell.coordinate)+':'+str(filename))
            except AttributeError:
                pass
            
    return set(referencesexcel),set(reflocations)
        
def processfactura(filename,referecelist,reflocations):
    def colorcell(color):
        if color=='red':
            color = openpyxl.styles.colors.Color(rgb='00FF0000')
        if color=='green':
            color = openpyxl.styles.colors.Color(rgb='00008000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color)
        return my_fill
    referencesonfacts=[]
    validsreferences=[]
    invalidscolumns=[]
    validscolumns=[]
    wb_obj = openpyxl.load_workbook(filename)
    ws = wb_obj.active
    
    for thenumber,row in enumerate(ws.iter_rows()):
        if platform.system() =="Windows":
            clear = lambda: os.system('cls')
            # clear()
        else:
            clear = lambda: os.system('clear')
            # clear()
        # if random.randint(1,10)>=8:
        progressbar(int(thenumber),int(ws.max_row))
        response=None
        for cell in row:
            if str((cell.value)) =='Referencia':
                reflocation=str((cell.coordinate)[0])
            try:
                if cell.column_letter==reflocation:
                    try:
                        if re.search('\d+',str((cell.value))):
                            extractedref=re.search('\d+',str((cell.value))).group()
                            if extractedref.startswith('000000'):
                                # print('detectado ref provincial 6 ceros')
                                extractedref=extractedref[6:]
                            if extractedref.startswith('0000'):
                                # print('detectado ref 4 ceros')
                                extractedref=extractedref[4:]
                            if extractedref.startswith('000'):
                                # print('detectado ref 3 ceros')
                                extractedref=extractedref[3:]
                            if extractedref.startswith('00'):
                                # print('detectado ref 2 ceros')
                                extractedref=extractedref[2:]
                            if extractedref.startswith('0'):
                                # print('detectado ref 1 cero')
                                extractedref=extractedref[1:]
                            # print(extractedref)
                            resx = difflib.get_close_matches(extractedref, referecelist,cutoff=0.82)
                            if len(resx)== 1:
                                for i in reflocations:
                                    refl=i.split(':')
                                    if(str(extractedref).find(str(refl[0]))==0):
                                        res=True
                                        thatlocation=i
                            if len(resx)>= 2:
                                if str((re.sub(r'[^0-9]', '', resx[0]))) != str((re.sub(r'[^0-9]', '', resx[1]))):
                                    if len(str((re.sub(r'[^0-9]', '', resx[0])))) == len(str((re.sub(r'[^0-9]', '', resx[1])))):
                                        for i in reflocations:
                                            refl=i.split(':')
                                            if(resx[0].find(refl[0])==0):
                                                thatlocation1=i
                                                refloc1=refl[1]
                                                filename1=refl[2]
                                            if(resx[1].find(refl[0])==0):
                                                thatlocation12=i
                                                refloc2=refl[1]
                                                filename2=refl[2]
                                        # print(resx)
                                        response=pyautogui.confirm(text='encontre una referencia muy parecida, la referencia es '+extractedref+'\nlas referencias encontradas fueron estas;\n'+str(resx[0])+' en '+str(filename1) +' en la locacion :'+str(refloc1) +' y \n'+str(resx[1])+ ' en '+str(filename2) +' en la locacion: '+str(refloc2)+'\nlas tomo validas ?', title='error', buttons=['Si', 'No'])
                                        if response =='Si':
                                            validscolumns.append(cell.coordinate)
                                        if response =='No':
                                            invalidscolumns.append(cell.coordinate)
                            if extractedref in referecelist:
                                # print(extractedref)
                                if response is None:
                                    validscolumns.append(cell.coordinate)
                                # print(reflocations[int(list(referecelist).index(str(extractedref)))])
                                # print(int(list(referecelist).index(str(extractedref))))
                                # print('encontre referencia valida')
                                
                            else:
                                if response is None:
                                    invalidscolumns.append(cell.coordinate)
                    except Exception as e: 
                        # print(e)
                        pass
                        
            except UnboundLocalError:
                pass
    
    for coordinates in invalidscolumns:
        # print(coordinates)
        ws[coordinates].fill = colorcell('red')
    for coordinates in validscolumns:
        ws[coordinates].fill = colorcell('green')
        # ws[coordinates].fill = 'greenFill'
    wb_obj.save(filename)  # save the workbook
    wb_obj.close()
    return str(len(validscolumns)), str(len(invalidscolumns))


for file in os.listdir():
    if re.findall('factura',str(((file).lower()))):
        if file.endswith('.xls'):
            x2x = XLS2XLSX(file)
            newfilename=str(file[:-3])+'xlsx'
            os.remove(file)
            wb = x2x.to_xlsx(newfilename)
            facturafilename=(newfilename)
            if re.findall('banesco',str(((file).lower()))):
                banescofiles.append(file)
            if re.findall('mercantil',str(((file).lower()))):
                mercantilfiles.append(file)
            if re.findall('bnc',str(((file).lower()))):
                bncfiles.append(file)
        if file.endswith('.xlsx'):
            facturafilename=file
            if re.findall('banesco',str(((file).lower()))):
                banescofiles.append(file)
            if re.findall('mercantil',str(((file).lower()))):
                mercantilfiles.append(file)
            if re.findall('bnc',str(((file).lower()))):
                bncfiles.append(file)
    else:
        if file.endswith('.pdf'):
            if re.findall('banesco',str(((file).lower()))):
                banescofiles.append(file)
            if re.findall('mercantil',str(((file).lower()))):
                mercantilfiles.append(file)
            if re.findall('bnc',str(((file).lower()))):
                bncfiles.append(file)
            gettedref,gettedlocations=processpdf(file)
            references.extend(gettedref)
            reflocations.extend(gettedlocations)
        if file.endswith('.xls'):
            x2x = XLS2XLSX(file)
            newfilename=str(file[:-3])+'xlsx'
            os.remove(file)
            wb = x2x.to_xlsx(newfilename)
            references.extend(processxls(newfilename))
            if re.findall('banesco',str(((file).lower()))):
                banescofiles.append(file)
            if re.findall('mercantil',str(((file).lower()))):
                mercantilfiles.append(file)
            if re.findall('bnc',str(((file).lower()))):
                bncfiles.append(file)
        if file.endswith('.xlsx'):
            gettedref,gettedlocations=processxls(file)
            references.extend(gettedref)
            reflocations.extend(gettedlocations)
            if re.findall('banesco',str(((file).lower()))):
                banescofiles.append(file)
            if re.findall('mercantil',str(((file).lower()))):
                mercantilfiles.append(file)
            if re.findall('bnc',str(((file).lower()))):
                bncfiles.append(file)
if len(bncfiles)==0:
    response=pyautogui.confirm(text='no se han encontrado los movimientos de bnc\ndesea continuar ?', title='error', buttons=['Si', 'No'])
    if response =='No':
        exit()
if len(banescofiles)==0:
    response=pyautogui.confirm(text='no se han encontrado los movimientos de banesco\ndesea continuar ?', title='error', buttons=['Si', 'No'])
    if response =='No':
        exit()
if len(mercantilfiles)==0:
    response=pyautogui.confirm(text='no se han encontrado los movimientos de mercantil\ndesea continuar ?', title='error', buttons=['Si', 'No'])
    if response =='No':
        exit()
if not facturafilename is None:
    valids,invalids=processfactura(facturafilename,set(references),reflocations)
    response=pyautogui.confirm(text='se han encontrado '+valids+' referencias validas en la factura\nse han encontrado '+invalids+' invalidas de '+str(int(valids)+int(invalids))+' en total', title='error', buttons=['Aceptar'])
